# -*- coding: utf-8 -*-
"""
Created on Fri Apr  9 15:06:32 2021

@author: Nicolas Spogis
"""

# importar os pacotes necessários
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import matplotlib.pyplot as plt
from wordcloud import WordCloud, STOPWORDS

global MinChar
MinChar = 2

def GetURLsFromDomain(domain_http):
    # Open The Output Excel
    if os.path.exists("./SiteLists/URL_List.xlsx"):
        os.remove("./SiteLists/URL_List.xlsx")
    
    wb = Workbook()
    wb.save(filename = './SiteLists/URL_List.xlsx')
    workbook = load_workbook(filename="./SiteLists/URL_List.xlsx")
    sheet = workbook.active
    
     
    url = domain_http
    reqs = requests.get(url)
    soup = BeautifulSoup(reqs.text, 'html.parser')
     
    for link in soup.find_all('a'):
        TempList = link.get('href')
        rows = sheet.max_row
        sheet.cell(row=rows+1, column=1).value = TempList
        
        
    workbook.save("./SiteLists/URL_List.xlsx")
    workbook.close()
    
def getSiteList(open_file):
    global SitesURLs
    open_file="./SiteLists/"+open_file
    workbook = load_workbook(open_file)
    sheet = workbook.active
    SitesURLs =[]

    for cell in sheet['A']:
        if cell.value is not None:
            SitesURLs.append(cell.value)
            
    workbook.close()

def getH1H2Data():
    
    # Open The Output Excel
    if os.path.exists("H1H2Data.xlsx"):
        os.remove("H1H2Data.xlsx")

    wb = Workbook()
    wb.save(filename = 'H1H2Data.xlsx')
    workbook = load_workbook(filename="H1H2Data.xlsx")
    sheet = workbook.active

    for url in SitesURLs:
        reqs = requests.get(url)
        soup = BeautifulSoup(reqs.text, 'lxml')
        for heading in soup.find_all(["h1", "h2"]):
            rows = sheet.max_row
            #sheet.cell(row=rows+1, column=1).value = (heading.name + ' ' + heading.text.strip())
            s = heading.text.strip()         
            words = s.split()
            for SingleWord in words:
                rows = sheet.max_row
                if len(SingleWord)>MinChar: 
                    sheet.cell(row=rows+1, column=1).value = SingleWord
    
    workbook.save("H1H2Data.xlsx")
    workbook.close()

def GenerateWordCloud():
    
    # lista de stopword
    STOPWORDS_DATA = []
    workbook = load_workbook(filename="./Others/STOPWORDS.xlsx")
    sheet = workbook.active
    for cell in sheet['A']:
        print(cell.value)
        STOPWORDS_DATA.append(cell.value)
        
    workbook.close()
    
    stopwords = set(STOPWORDS)
    stopwords.update(STOPWORDS_DATA)

    # Start by opening the spreadsheet and selecting the main sheet
    workbook = load_workbook(filename="H1H2Data.xlsx")
    sheet = workbook.active
    summary =[]
    
    for cell in sheet['A']:
        summary.append(cell.value)
    
    # concatenar as palavras
    all_summary = ' '.join([str(elem) for elem in summary])
    
    
    # gerar uma wordcloud
    wordcloud = WordCloud(stopwords=stopwords,
                          background_color="white",
                          width=1000, height=1000, max_words=300,
                          max_font_size=200,
                          min_font_size=1).generate(all_summary)
     
    # mostrar a imagem final
    fig, ax = plt.subplots(figsize=(100,100))
    ax.imshow(wordcloud, interpolation='bilinear')
    ax.set_axis_off()
     
    plt.imshow(wordcloud);
    
    # Determine incremented filename
    filename = "./Pictures/WordCloud.png"
    wordcloud.to_file(filename)

#GetURLsFromDomain("https://www.ingredion.com/sa/pt-br.html")
getSiteList("FoodIngredients.xlsx")
getH1H2Data()
GenerateWordCloud()
