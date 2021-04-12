# -*- coding: utf-8 -*-
"""
Created on Fri Apr  9 17:43:21 2021

@author: nicol
"""
# importar os pacotes necessários
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from GoogleNews import GoogleNews

global links

googlenews = GoogleNews()

# Setup the research
keywords="software juridico"
period='60d'

googlenews.clear()
googlenews.set_lang('pt')
googlenews.set_period(period)
googlenews.search(keywords)
googlenews.get_page(2)
aaa=googlenews.total_count()

links = googlenews.get_links()


def getNewsLinks():
    # Open The Output Excel
    if os.path.exists("./NewsLinks/NewsLinks.xlsx"):
        os.remove("./NewsLinks/NewsLinks.xlsx")

    wb = Workbook()
    wb.save(filename = './NewsLinks/NewsLinks.xlsx')
    workbook = load_workbook(filename="./NewsLinks/NewsLinks.xlsx")
    sheet = workbook.active

    for url in links:
        rows = sheet.max_row
        sheet.cell(row=rows+1, column=1).value = url
    
    workbook.save("./NewsLinks/NewsLinks.xlsx")
    workbook.close()

getNewsLinks()